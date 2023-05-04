# -*- coding: utf-8 -*-
"""
Created on Tue May  2 16:09:14 2023

@author: Filip Poposki
"""

import os
from openpyxl import Workbook
import pandas as pd

def get_file_list():
    '''
    Returns
    -------
    list
        list contains names of files found in folder. Alarms user if no files found.
    '''
    if len(os.listdir()) == 1:
        print(f"Working directory: {os.getcwd()}")
        print(f"Found the following files: {os.listdir()}\n")
    else:
        print(f"Working directory: {os.getcwd()}")
        print(f"Found the following files: {os.listdir()}\n")
    return os.listdir()


def check_count_of_files(list_files):
    '''
    Parameters
    ----------
    list_files : list
        list of files found in folder.
        
    Returns
    -------
    bool
        pass if number of files is ok to proceed further

    '''
    if len(list_files) < 3:
        print('Not enough total files in folder')
        return False
    else: 
        return True
    
    
def find_xlsx_file(list_files_folder):
    '''
    Parameters
    ----------
    list_files : list
        list of files found in folder.
        
    Returns
    -------
    list
        returns list of files where .xlsx is found
    '''
    xlsx_files_list = []
    for i in range(0,len(list_files_folder)):
        if list_files_folder[i].endswith(".xlsx"):
            xlsx_files_list.append(list_files_folder[i])
            
    return xlsx_files_list

def check_count_xslx_files(list_xlsx_files):
    '''
    Parameters
    ----------
    list_files : list
        list of files with .xlsx found from find_xlsx_file()
        
    Returns
    -------
    bool
        pass if number of files is ok to proceed further
    '''
    if len(list_xlsx_files) == 2:
        return True
    else:
        print("Number of .xlsx files is not adequate")
        return False

def bom_to_grouped_df(bom_filename):
    '''
    Parameters
    ----------
    bom_filename : string
        DESCRIPTION.

    Returns
    -------
    grouped_df : dataframe
        DESCRIPTION.

    '''    
    # Import file as Dataframe
    df = pd.read_excel(f"{bom_filename}")
    
    #Drop NaN rows
    df = df.dropna(axis = 0)
    
    #Drop first rows (not needed)
    df = df.iloc[3:]
    
    #Change collumn name
    df.columns.values[0] = "ID"
    
    #Leave only rows that dont contain the string "Level" 
    df = df.loc[~df['ID'].str.contains('Level')]
    
    #Replace specific strings that are not needed 
    #keep in mind the number of empty spaces!!!
    # count empty spaces in excel and then insert the exact number
    df['ID'] = df['ID'].str.replace('LcsReleased     ', '')
    df['ID'] = df['ID'].str.replace('LcsWorking     ', '')
    df['ID'] = df['ID'].str.replace('LcsUpgrade      ', '')
    
    
    # Leave only rows that contain the string "PC"
    df = df[df['ID'].str.contains('PC')]
    
    # Split collumn ID into 2 collumns, first is Part number with split_length of chars, and the second is the rest
    split_length = 8
    df = pd.DataFrame({'Part_number': df['ID'].str[:split_length], 'Ref': df['ID'].str[split_length:]})
    
    #Split Ref collumn to have only the designator (found after string PC)
    df['Ref'] = df['Ref'].str.split('PC').str[-1]

    # Remove empty spaces from ref column
    df['Ref'] = df['Ref'].str.replace(" ","")
    
    
    #Grouping of df by ref/designator (that is unique, can have multiple part numbers on a single designator)
    # values are given in list
    grouped_df = df.groupby('Ref').agg({'Part_number': list})
    
    return grouped_df



def get_dict_from_processed_bom_df(processed_df_from_bom):
    '''
    Parameters
    ----------
    processed_df_from_bom : Dataframe
        DESCRIPTION.

    Returns
    -------
    bom_dict : dict
        DESCRIPTION.

    '''    
    # Creates a dict with unique Ref,list
    bom_dict = processed_df_from_bom.to_dict()['Part_number']
    
    # Delete all values with no key (key = ''), meaning with no designator
    del bom_dict[""]
    
    return bom_dict

    

def create_bom_difference_dict(dict1,dict2,filename1,filename2):
    '''
    Parameters
    ----------
    dict1 : dictionary
        DESCRIPTION.
    dict2 : dictionary
        DESCRIPTION.
    filename1 : string
        DESCRIPTION.
    filename2 : string
        DESCRIPTION.

    Returns
    -------
    None.
    '''
    differences = {}
    
    for key in dict1:
        if key not in dict2:
            differences[key] = (dict1[key], "Not populated in this BOM")
        elif key in dict2:
            if dict1[key] != dict2[key]:
                differences[key] = (dict1[key], dict2[key])
            
    for key in dict2:
        if key not in dict1:
            differences[key] = ("Not populated in this BOM", dict2[key])
        elif key in dict1:
            if dict1[key] != dict2[key]:
                differences[key] = (dict1[key], dict2[key])

    return differences
    
    
def generate_bom_difference_report(dict1,dict2,dict_differences,bom_name1,bom_name2):
    '''
    
    Parameters
    ----------
    dict1 : dictionary
        dictionary contains ref designators and PNs from bom_name1.
    dict2 : dictionary
        dictionary contains ref designators and PNs from bom_name2.
    dict_differences : dictionary
        contains dictionary of differences.
    bom_name1 : string
        name of first file.
    bom_name2 : string
        name of second file.

    Returns
    -------
    None.

    '''


    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Comparison of BOMs"
    
    sheet1.append(["Reference Designator", bom_name1, bom_name2])
       
    #Populates the comparison sheet with the values from the differences list
    for key in dict_differences:
        sheet1.append([f"{key}", f"{dict_differences[key][0]}", f"{dict_differences[key][1]}"])

    #Creates workbook to have overview of ref designators and PNs in bom_name1
    sheet2 = workbook.create_sheet(f"{bom_name1}")
    sheet2.append(["Reference Designator", bom_name1])
    
    for key in dict1:
        sheet2.append([f"{key}", f"{dict1[key]}"])
               
    #Creates workbook to have overview of ref designators and PNs in bom_name2
    sheet3 = workbook.create_sheet(f"{bom_name2}")
    sheet3.append(["Reference Designator", bom_name2])

    for key in dict2:
        sheet3.append([f"{key}", f"{dict2[key]}"])                  
    
    workbook.save(f"Comparison results for BOMs {bom_name1} and {bom_name2}.xlsx")
        
    print("Report generated. Close program and check the excel file.")
     



if __name__ == "__main__":
    
    
    files_in_folder = get_file_list()
    
         
    if check_count_of_files(files_in_folder):
        
    
        xlsx_files_in_folder = find_xlsx_file(files_in_folder)
        
        
        if check_count_xslx_files(xlsx_files_in_folder):
            
            bom_filename1 = xlsx_files_in_folder[0]
            bom_filename2 = xlsx_files_in_folder[1] # Useful for further processing
            
            bom_df1 = bom_to_grouped_df(bom_filename1)
            bom_df2 = bom_to_grouped_df(bom_filename2)
            
            bom_dict1 = get_dict_from_processed_bom_df(bom_df1)
            bom_dict2 = get_dict_from_processed_bom_df(bom_df2)
            
            
            bom_differences_dict = create_bom_difference_dict(bom_dict1, bom_dict2, bom_filename1, bom_filename2)
            
            generate_bom_difference_report(bom_dict1, bom_dict2, bom_differences_dict, bom_filename1, bom_filename2)
            

os.system("pause")

